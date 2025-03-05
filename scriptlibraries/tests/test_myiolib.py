import unittest as ut
import os
import sys
import inspect
import tempfile
import csv
import shutil
import openpyxl

currentdir = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
parentdir = os.path.dirname(currentdir)
sys.path.insert(0, parentdir) 

import myiolib as lib 

class TestMyiolibpy(ut.TestCase):
    '''Tests everything except file writing in myiolib.py'''

    def test_load_config(self): # The value of this test depends on dumyconfig.yaml
        '''Tests load_config function with a single dummy file'''
        self.assertEqual(lib.load_config('dummyconfig.yaml','scriptlibraries\\tests'),('2','1','3','4',5,6,7,False,True))

    def test_resolve_file_name(self): # Many tests are windows specific.
        '''Tests simple joining, empty subfolder and file expection.'''
        self.assertEqual(lib.resolve_file_name('A','B'),"B\\A")
        self.assertEqual(lib.resolve_file_name('the_underscored','underscored_too'),'underscored_too\\the_underscored')
        self.assertEqual(lib.resolve_file_name('file',''),'file')
        with self.assertRaises(NameError):
            lib.resolve_file_name('','subfolder')

    def test_read_pdf(self): # Whenever you update the pdf files in tests file also update test_read_all_pdfs 
        '''Tests read_pdf function with multipage pdf, cyrilic pdf, pdfs with spaces in their names and a nonexisting pdf'''
        self.assertEqual(lib.read_pdf('sample.pdf','scriptlibraries\\tests'),'sample! ')
        self.assertEqual(lib.read_pdf('sample_cyrilic_content.pdf','scriptlibraries\\tests'),'Ћирлилица бајо :0 ')
        self.assertEqual(lib.read_pdf('Ћирличнинасловуву.pdf','scriptlibraries\\tests'),'Osvete celave latinice ')
        self.assertEqual(lib.read_pdf('razmaknuti test.pdf','scriptlibraries\\tests'),'razmaknuti ')
        with self.assertRaises(FileNotFoundError):
            lib.read_pdf('nepostojeci.pdf')
        self.assertEqual(lib.read_pdf('multipage.pdf','scriptlibraries\\tests'),'strana1 \n  strana2 ')

    def test_read_all_pdfs(self): # If you add any files at all to the tests folder you must update this pdf!
        '''Tests if read_all_pdfs actually reads all the pdfs through a single test of the entire folder tests'''
        self.assertEqual(lib.read_all_pdfs('scriptlibraries\\tests'),[None, 'strana1 \n  strana2 ', 'razmaknuti ', 'sample! ', 'Ћирлилица бајо :0 ', None, None, 'Osvete celave latinice '])

class TestCSVExport(ut.TestCase):

    def setUp(self):
        self.test_base_dir = tempfile.mkdtemp() # Temporary directory for testing
        self.outputs_dir = os.path.join(self.test_base_dir, 'outputs') # Adding temporary outputs subdirectory
        os.makedirs(self.outputs_dir, exist_ok=True)        
        
        self.original_path = os.getcwd() # Save original sys.path for restoring after the test
        os.chdir(self.test_base_dir) # Modify sys.path to ensure the test directory is used for file resolution
        
        self.sample_data = [
            ['Dimitrije', '21', 'Software Engineer'],
            ['Bratislav', '51', 'Project engineer'],
            ['Vikica', '50', 'Student services']
        ]
        self.column_names = ['Name', 'Age', 'Profession']

    def test_export_to_csv_write_mode(self):
        """Test exporting data in write mode"""
        file_name = 'test_write.csv'

        lib.export_to_csv( #Export data
            data=self.sample_data, 
            column_names=self.column_names,
            file_name=file_name,
            subfolder='outputs',
            mode='w'
        )
        
        full_path = os.path.join(self.outputs_dir, file_name) # Full path to the exported file
         
        self.assertTrue(os.path.exists(full_path), f"File {full_path} was not created")

        with open(full_path, 'r', newline='') as csvfile: # Read and verify the contents
            reader = csv.reader(csvfile)
            rows = list(reader)
            
            self.assertEqual(rows[0], self.column_names, "Headers do not match")

            for i, row in enumerate(rows[1:], start=0):
                self.assertEqual(row, self.sample_data[i], f"Data row {i} does not match")

    def test_export_to_csv_append_mode(self):
        """Test appending data to an existing CSV file"""
        file_name = 'test_append.csv'
        
        # First export - write mode, so the tests don't depend on the previous test
        lib.export_to_csv(
            data=self.sample_data, 
            column_names=self.column_names,
            file_name=file_name,
            subfolder='outputs',
            mode='w'
        )
        
        # Additional data to append
        additional_data = [
            ['Lenka', '17', 'Student'],
            ['Charlie', '40', 'Director']
        ]
        
        lib.export_to_csv( # Apeend data
            data=additional_data, 
            file_name=file_name,
            subfolder='outputs',
            mode='a'
        )
        
        full_path = os.path.join(self.outputs_dir, file_name) # Full path to the exported file
        with open(full_path, 'r', newline='') as csvfile: # Read and verify the contents
            reader = csv.reader(csvfile)
            rows = list(reader)
            
            self.assertEqual(rows[0], self.column_names, "Headers do not match")
            
            for i, row in enumerate(rows[1:4], start=0):
                self.assertEqual(row, self.sample_data[i], f"Original data row {i} does not match")
            
            for i, row in enumerate(rows[4:], start=0):
                self.assertEqual(row, additional_data[i], f"Appended data row {i} does not match")

    def test_export_to_csv_empty_data(self):
        """Test exporting an empty dataset"""
        file_name = 'test_empty.csv'
        
        # Export empty data
        lib.export_to_csv(
            data=[], 
            column_names=self.column_names,
            file_name=file_name,
            subfolder='outputs'
        )
        
        full_path = os.path.join(self.outputs_dir, file_name)
        self.assertTrue(os.path.exists(full_path), f"File {full_path} was not created")
        
        # Read and verify the contents
        with open(full_path, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            rows = list(reader)
            self.assertEqual(len(rows), 1, "Empty data file should only contain headers")
            self.assertEqual(rows[0], self.column_names, "Headers do not match")

    def tearDown(self):
        os.chdir(self.original_path) # Change back to the original directory
        shutil.rmtree(self.test_base_dir) # Remove the temporary directory

class TestExcelExport(ut.TestCase):
    def setUp(self):
        # Create a temporary base directory for all test files
        self.test_base_dir = tempfile.mkdtemp()
        
        # Create outputs directory within the base directory
        self.outputs_dir = os.path.join(self.test_base_dir, 'outputs')
        os.makedirs(self.outputs_dir, exist_ok=True)
        
        # Modify sys.path to ensure the test directory is used for file resolution
        self.original_path = os.getcwd()
        os.chdir(self.test_base_dir)
        
        # Sample data for testing
        self.sample_data = [
            ['Name', 'Age', 'Profession'],
            ['John', '30', 'Engineer'],
            ['Jane', '28', 'Designer'],
            ['Bob', '35', 'Manager']
        ]

    def test_export_to_xlsx_basic(self):
        """Test basic Excel file export"""
        file_name = 'test_basic.xlsx'
        
        # Export the data
        lib.export_to_xslsx(
            data=self.sample_data, 
            file_name=file_name,
            subfolder='outputs'
        )
        
        # Full path to the exported file
        full_path = os.path.join(self.outputs_dir, file_name)
        
        # Verify the file was created
        self.assertTrue(os.path.exists(full_path), f"File {full_path} was not created")
        
        # Read and verify the contents
        wb = openpyxl.load_workbook(full_path)
        ws = wb.active
        
        # Check data rows
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=0):
            self.assertEqual(list(row), self.sample_data[row_idx], 
                             f"Row {row_idx} does not match expected data")

    def test_export_to_xlsx_bold_header(self):
        """Test Excel export with bold header"""
        file_name = 'test_bold_header.xlsx'
        
        # Export the data with bold header
        lib.export_to_xslsx(
            data=self.sample_data, 
            file_name=file_name,
            subfolder='outputs',
            bold_header=True
        )
        
        # Full path to the exported file
        full_path = os.path.join(self.outputs_dir, file_name)
        
        # Verify the file was created
        self.assertTrue(os.path.exists(full_path), f"File {full_path} was not created")
        
        # Read and verify the contents
        wb = openpyxl.load_workbook(full_path)
        ws = wb.active
        
        # Check header row formatting
        header_row = list(ws.iter_rows())[0]
        for cell in header_row:
            self.assertTrue(cell.font.bold, "Header cell should be bold")
        
        # Check data content
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=0):
            self.assertEqual(list(row), self.sample_data[row_idx], 
                             f"Row {row_idx} does not match expected data")

    def test_export_to_xlsx_empty_data(self):
        """Test exporting an empty dataset"""
        file_name = 'test_empty.xlsx'
        
        # Export empty data
        lib.export_to_xslsx(
            data=[], 
            file_name=file_name,
            subfolder='outputs'
        )
        
        # Full path to the exported file
        full_path = os.path.join(self.outputs_dir, file_name)
        
        # Verify the file was created
        self.assertTrue(os.path.exists(full_path), f"File {full_path} was not created")
        
        # Read and verify the contents
        wb = openpyxl.load_workbook(full_path)
        ws = wb.active
        
        # Check that no rows are present
        rows = list(ws.iter_rows(values_only=True))
        self.assertEqual(len(rows), 0, "Empty data should result in an empty worksheet")

    def test_export_to_xlsx_nonexistent_directory(self):
        """Test exporting to a non-existent directory"""
        # Remove the outputs directory to simulate non-existent directory
        shutil.rmtree(self.outputs_dir)
        
        file_name = 'test_new_dir.xlsx'
        
        # Export data (should create the directory)
        lib.export_to_xslsx(
            data=self.sample_data, 
            file_name=file_name,
            subfolder='outputs'
        )
        
        # Verify the directory was created
        self.assertTrue(os.path.exists(os.path.join(self.test_base_dir, 'outputs')), 
                        "Directory should be created if it doesn't exist")
        
        # Verify the file was created
        full_path = os.path.join(self.test_base_dir, 'outputs', file_name)
        self.assertTrue(os.path.exists(full_path), f"File {full_path} was not created")
        
        # Read and verify the contents
        wb = openpyxl.load_workbook(full_path)
        ws = wb.active
        
        # Check data rows
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=0):
            self.assertEqual(list(row), self.sample_data[row_idx], 
                             f"Row {row_idx} does not match expected data")

    def tearDown(self):
        # Change back to the original directory
        os.chdir(self.original_path)
        
        # Remove the temporary directory
        shutil.rmtree(self.test_base_dir)

if __name__=='__main__':
    ut.main()